# MIT License
#
# Copyright (c) 2019 Loh Yu Chen & Chi Junxiang
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.
# ================================================================================

"""Provides utility functions.

This module contains utility functions for main.py and analyse.py.
Those functions are placed into this module to prevent spaghetti code in
both of those scripts.
"""
# Imports
import os
import pickle
from random import choices, choice
from string import ascii_letters, digits
from csv import reader

import plotly
from flask import Markup
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

global COLORS
COLORS = ["#4e73df", "#6610f2", "#6f42c1", "#e83e8c", "#e74a3b", "#fd7e14", "#f6c23e", "#1cc88a", "#20c9a6",
          "#36b9cc"]
COLORS = ["#0c1941", "#183281", "#2043ac", "#2854d7", "#4e73df"]


# Misc
def secure(length):
    """Returns a random string of length

    Args:
        length(int): Length of the result

    Returns:
        Random string of length

    >>> len(secure(16)) == 16
    True
    """
    return "".join(choices(ascii_letters + digits, k=length))


# File Utils
def parse_csv(csv_file):
    """Parses an csv file by column

    Parse the csv file, which has the responses to a survey,
    represented by csv_file

    For example:
    Do you like python?,What other languages do you use?
    Yes,Go
    No,Java
    Yes,C++

    Args:
        csv_file(str): The file name of the excel_file to be parsed

    Returns:
        A dictionary mapping the question to a tuple of responses
        For example:
        {"Do you like Python?": ("Yes","No","Yes"),
        "What other languages do you use?": ("Go","Java","C++")}
    """
    with open(csv_file, "r") as f:
        rows = list(reader(f, delimiter=','))
        responses = []
        for i, header in enumerate(rows[0]):
            responses.append((header, [row[i] for row in rows[1:]]))
    return dict(responses)


def parse_excel(excel_file):
    """Parses an Excel file by column

    Parse the Excel file, which has the responses to a survey,
    represented by excel_file, by column

    For example:
     |        A          |
    1|Do you like python?|
    2|Yes                |
    3|No                 |
    4|Yes                |

    Args:
        excel_file(str): The file name of the excel_file to be parsed

    Returns:
        A dictionary mapping the question to a tuple of responses
        For example:
        {"Do you like Python?": ("Yes", "No", "Yes")}

    Raises:
        KeyError: Question not present in column {i+1}
    """
    wb = load_workbook(excel_file)
    ws = wb.active
    cell_values = []
    for i, col in enumerate(ws.columns):
        if col[0].value:
            cell_values.append(
                [str(cell.value) for cell in col if cell.value is not None]
            )
        else:
            raise KeyError(f"Question not present in column {get_column_letter(i + 1)}")
    return dict([(col[0], col[1:]) for col in cell_values])


def parse_config(config_file):
    """Parses a config file

    Parses a config file represented by config_file. Config file
    contains the data-type of the responses.

    For example:
    1 categorical
    2 numerical
    3 ignore
    4 openended

    Args:
        config_file(str): The filename of the config file to be parsed

    Returns:
        A dictionary mapping the datatype to the question number
        For example:
        {1: "categorical", 2: "numerical", 3: "ignore", 4: "openended", 5: "multicategorical"}

    Raises:
        ValueError: Data-type not supported
    """
    with open(config_file, mode="r", encoding="utf-8") as f:
        lines = [line.lower() for line in f.read().split("\n") if line != ""]
        lines = [line.split(" ") for line in lines]
        accepted = ("ignore", "numerical", "categorical", "openended", "multicategorical")
        for i in lines:
            if not i[0].isdigit() or i[1] not in accepted:
                raise TypeError(
                    f"Line'{i}': parse_config only accepts lines with format <qn_no> <qn_type>"
                )
        lines = [(int(line[0]), line[1]) for line in lines]
        return dict(lines)


def to_config(directory, config):
    """"Creates/modifies a config file

    Creates/modifies a config file based on config given
    If config_file exists, it will modify the given config_file;
    else, it will create a new config file

    Args:
        directory(str) : Directory to save config file in
        config(dict) : Dictionary mapping datatype to question number
    Returns:
        String containing the path to the config file
    """
    to_write = [f"{i[0]} {i[1]}\n" for i in config.items()]
    file_path = os.path.join(directory, "config_file.txt")
    with open(file_path, "w") as config_f:
        config_f.writelines(to_write)
    return file_path


# Prediction
class Predictor(object):
    """TextBlob classifier that predicts if qn is either
    categorical, numerical, or openended.
    """

    def __init__(self):
        """Loads pre-trained TextBlob classifier."""
        with open("model.pickle", "rb") as pick:
            self.classifier = pickle.load(pick)

    def predict(self, qns):
        """Predicts if qns in qns are either categorical, numerical, or openended."""
        return [self.classifier.classify(qn) for qn in qns]


# Plotting
def pie(title, labels, values, hole=0.4):
    """Creates a pie chart

    Parses data passed in and creates a pie chart.

    Args:
        title(str): Title of the pie chart
        labels(list): Labels of each slice in the pie chart
        values(list): Value of each slice
        hole(float): Size of the hole in the pie chart

    Returns:
        HTML div of the pie chart
    """
    fig = plotly.graph_objs.Figure(
        data=[
            plotly.graph_objs.Pie(
                labels=labels,
                values=values,
                hole=hole,
                hoverinfo="label+percent",
                text=labels,
                marker=dict(colors=COLORS, line=dict(color="#FFFFFF", width=2)),
                sort=False,
                showlegend=False,
            )
        ]
    )

    fig.layout.font = dict(family="Nunito", size=18, color="#858796")

    return Markup(plotly.offline.plot(fig, include_plotlyjs=False, output_type="div"))


def chunk(iterable, size):
    """Splits iterable into chunks

    Splits iterable into multiple chunks

    Args:
        iterable(iterable): Iterable to be split
        size(int): Size of each chunk

    Yields:
        One chunk

    >>> list(chunk([1, 2, 3, 4], 2))
    [[1, 2], [3, 4]]
    """
    for i in range(0, len(iterable), size):
        yield iterable[i: (i + size)]


def random_colour(word, font_size, position, orientation, font_path, random_state):
    return choice(COLORS)


if __name__ == "__main__":
    parse_csv()
    # pass  # Testing is over :D
